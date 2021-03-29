# -*- coding: utf-8 -*-
import utility
import unmask
import openpyxl

EMILPO_KEY_DATE = "2-5-2021"

def run_export_jobs(
    EXPORT_F2S,
    EXPORT_UNMATCHED,
    UPDATE_CONNECTIONS,
    EXPORT_UNMASKED,
    EXPORT_CMD_SPECS,
    EXPORT_UNMASKED_CMD_SPECS,
    COMMAND_EXPORT_LIST,
    face_space_match,
    all_faces_to_matched_spaces,
    attached_faces_to_matched_spaces,
    cmd_metrics,
    ar_cmd_metrics,
    ac_ar_metrics,
    curorg_metrics,
    unmatched_faces,
    drrsa,
    address_data,
    acronym_list,
    AC_KEY_FILE,
    RC_KEY_FILE
):
    print(" - Running export jobs")
    if(EXPORT_F2S):
        print("  - Exporting all_faces_to_matched_spaces and command metrics")
        export_matches(face_space_match, all_faces_to_matched_spaces)
        export_metrics(cmd_metrics, ar_cmd_metrics, ac_ar_metrics, curorg_metrics)
        
    if(EXPORT_UNMATCHED):
        print("  - Exporting unmatched faces")
        export_unmatched(unmatched_faces)
        
    if(UPDATE_CONNECTIONS):
        print("  - Updating Access DB connection files")
        update_connections(
            all_faces_to_matched_spaces, cmd_metrics, ar_cmd_metrics, curorg_metrics
        )
        
    if(EXPORT_UNMASKED):
        print("  - Exporting unmasked files to local directory")
        unmask.unmask_and_export(
            all_faces_to_matched_spaces,
            attached_faces_to_matched_spaces,
            utility.get_file_timestamp(),
            AC_KEY_FILE,
            RC_KEY_FILE
        )
    
    if(EXPORT_CMD_SPECS or EXPORT_UNMASKED_CMD_SPECS):
        print("  - Exporting command metrics workbooks")
        import analytics.cmd_metrics_package #Uncomment for debugging
        analytics.cmd_metrics_package.create_cmd_metrics_packages(
            all_faces_to_matched_spaces,
            drrsa,
            address_data,
            acronym_list,
            curorg_metrics,
            ar_cmd_metrics,
            unmask_ssn = EXPORT_UNMASKED_CMD_SPECS,
            date_time_string = utility.get_file_timestamp(),
            commands = COMMAND_EXPORT_LIST,
            emilpo_key = AC_KEY_FILE,
            tapdbr_key = RC_KEY_FILE
        )

def export_matches(face_space_match, all_faces_to_matched_spaces):
    face_space_match.to_csv(
        "..\export\\face_space_matches" 
        + utility.get_file_timestamp() 
        + ".csv"
    )
    all_faces_to_matched_spaces.to_csv(
        "..\export\\all_faces_to_matched_spaces" 
        + utility.get_file_timestamp() 
        + ".csv"
    )
    
def export_metrics(cmd_metrics, ar_cmd_metrics, ac_ar_metrics, curorg_metrics):
    cmd_metrics.to_csv(
        "../export/command_metrics"
        + utility.get_file_timestamp()
        + ".csv"
    )                
    ar_cmd_metrics.to_csv(
        "../export/ar_command_metrics"
        + utility.get_file_timestamp()
        + ".csv"
    )
    ac_ar_metrics.to_csv(
        "../export/ac_ar_command_metrics"
        + utility.get_file_timestamp()
        + ".csv"
    )
    curorg_metrics.to_csv(
        "../export/ar_curorg_rcc_metrics"
        + utility.get_file_timestamp()
        + ".csv"   
    )
    
def export_unmatched(unmatched_faces):
    unmatched_faces.to_csv(
        "..\export\\unmatched_faces" 
        + utility.get_file_timestamp() 
        + ".csv"
    )
    
def update_connections(all_faces_to_matched_spaces, cmd_metrics, ar_cmd_metrics, curorg_metrics):
    all_faces_to_matched_spaces.drop_duplicates(subset = "SSN_MASK").to_csv(
        "..\export\\for_connections\\all_faces_to_matched_space_latest.csv"
    )
    cmd_metrics.to_csv("..\export\\for_connections\\cmd_metrics.csv")
    ar_cmd_metrics.to_csv("..\export\\for_connections\\ar_cmd_metrics.csv")
    
    print("  - Generating new excel dashboards in exports folder")
    ac_cmd_metric_dashboard = openpyxl.load_workbook(
        filename = "F:/aos/f2s_project/export/ac_cmd_metric_dashboard_template.xlsx",
        data_only = False
    )
    del ac_cmd_metric_dashboard["CMD_Metrics"]
    ac_cmd_metric_dashboard.create_sheet("CMD_Metrics")
    
    for row in cmd_metrics[[
        "STRUC_CMD_CD", "ASSIGNED", "UNMATCHED", "MATCHED", "PERCENT_MATCHED",
        "UNMATCHED_NO_TEMPLET", "UNMATCHED_NO_UIC", "UICS_NOT_IN_AOS",
        "MATCHED_ASG_OLDER_THAN_POS", "METRIC_DATE", "COMPONENT"        
    ]].itertuples():
        ac_cmd_metric_dashboard["CMD_Metrics"].append(row)
        
    ac_cmd_metric_dashboard.save(
        "F:/aos/f2s_project/export/ac_cmd_metric_dashboard" 
        + utility.get_file_timestamp()
        + ".xlsx"
    )
    
    ar_cmd_metric_dashboard = openpyxl.load_workbook(
        filename = "F:/aos/f2s_project/export/ar_cmd_metric_dashboard_template.xlsx",
        data_only = False
    )
    del ar_cmd_metric_dashboard["CMD_Metrics"]
    ar_cmd_metric_dashboard.create_sheet("CMD_Metrics")
    
    del ar_cmd_metric_dashboard["CURORG_Metrics"]
    ar_cmd_metric_dashboard.create_sheet("CURORG_Metrics")
    
    del ar_cmd_metric_dashboard["AC_CMD_Metrics"]
    ar_cmd_metric_dashboard.create_sheet("AC_CMD_Metrics") 
    
    for row in ar_cmd_metrics.itertuples():
        ar_cmd_metric_dashboard["CMD_Metrics"].append(row)
    
    for row in curorg_metrics.itertuples():
        ar_cmd_metric_dashboard["CURORG_Metrics"].append(row)
        
    for row in cmd_metrics.itertuples():
        ar_cmd_metric_dashboard["AC_CMD_Metrics"].append(row)
    
    ar_cmd_metric_dashboard.save(
        "F:/aos/f2s_project/export/ar_cmd_metric_dashboard" 
        + utility.get_file_timestamp()
        + ".xlsx"
    )
    
    